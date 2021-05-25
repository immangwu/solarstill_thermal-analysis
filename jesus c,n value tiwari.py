####----------------------------------------------------------------------------------------------------------------------------------------######
                                                        ##Radiation Related Values##
#Conventions taken from G N Tiwari boobk
#vARIATION OF ATTENUATION FACTOR WITH WATER DEPTH PG NO 281 IN TIWARI BOOK


def latentheatofvapor(Temp):
    L=(2.4935*10**6)*(1-((9.4779*10**-4)*Temp)+((1.3132*10**-7)*Temp**2)-((4.7974*10**-9)*Temp**3));
    return L

#the values of Pw and Pg
import math
import numpy as np
def PT(Temp):
    PT = math.exp(25.317-(5144/(float(Temp)+273)));
    return PT
def TPP(Temp,T_w,T_g,K_i,d_f ,m_ew): #thermo physical properties
    density = 353.44/(Temp+273.15);  #Toyama etal 1987
    SHC = 999.2+(0.1434*Temp)+((1.101*10**-4)*Temp**2)-((6.7581*10**-8)*Temp**3); # specific heat capacity
    TC_K=0.0244+(0.7673*10**-4)*Temp # thermal Conductivity
    viscos=(1.718*10**-5)+(4.620*10**-8)*Temp
    diffus=(7.7255*10**-10) * (Temp+2730)**1.83
    exp_fac=1/(Temp+273.15) #beta
    L_v=latentheatofvapor(T_w);
    g=9.81
    P_w=PT(T_w);
    P_g=PT(T_g);
    delT=(T_w-T_g)+(((P_w-P_g)*(T_w+273))/((2.689*10**5)-P_w)); #https://sci-hub.se/https://www.sciencedirect.com/science/article/abs/pii/S2352152X20309543
    R=0.0163*(P_w-P_g)*(TC_K/d_f)*(3600/L_v)
    Gr=(g*exp_fac*(density**2)*(d_f **3)*delT)/(viscos**2)
    Pr= (viscos*SHC)/TC_K
    y=math.log(m_ew/R);x=math.log(Gr*Pr)
    return density,SHC,TC_K,viscos,diffus,exp_fac,L_v,P_w,P_g,delT,R,Gr,Pr,x,y
def _sum(arr):
    sum=0
    for i in arr:
        sum = sum + i
          
    return(sum) 





####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                            ###General variables###
depthw = 0.10; #depth of water in m 10 cm = 0.01m
alpha_w = 0.2 ; #water absorbtivity 0.2 from tiwari boobk
alpha_g = 0.0 ; # absorbtivity of glass cover
alpha_b = 0.075; # absorbtivity of basin cover
R_g = 0.05; # reflectivity of glass cover
R_w =0.05;  # reflectivity of water
eps_g = 0.94; #emisivity of glass cover
eps_w = 0.95 #emissivity of water
V=3;
V_basin =0; #basin wind velocity m/s zero always
h_wbslc=100; #W/m^2deg celcius taken from book need to find it heat transfer between the basin liner and the water
K_i = 0.04; #Insulation thermal conductivity [W/m K]
L_i = 0.5*10**-2; #Insulation thickness [m]
A_s=1;#surface area of the basin
d_f = 0.155 #Average space between water surface and glass cover)
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                        ##Excel##
import openpyxl
import xlrd
from pathlib import Path
import xlwt
from openpyxl import load_workbook

import math
loc = ("cn.xlsx");
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
wb2 = load_workbook (loc)
sheets=wb2.sheetnames
Sheet1=wb2 [sheets[0]]
sx=[]
sy=[]
sxy=[]
sx2=[]
for j in range(sheet.nrows-1):#sheet.nrows-1):
    Ta=float((sheet.cell_value(j+1, 0)))
    T_w=(sheet.cell_value(j+1, 1))
    T_g=(sheet.cell_value(j+1, 2))
    Temp=(T_w+T_g)/2#float((sheet.cell_value(j+1, 3)))
    
    m_ew=float((sheet.cell_value(j+1, 4)))
    density,SHC,TC_K,viscos,diffus,exp_fac,L_v,P_w,P_g,delT,R,Gr,Pr,x,y = TPP(Temp,T_w,T_g,K_i,d_f ,m_ew)
    
    Sheet1.cell(row=j+2,column=6).value =L_v
    Sheet1.cell(row=j+2,column=7).value =P_g
    Sheet1.cell(row=j+2,column=8).value =P_w
    Sheet1.cell(row=j+2,column=9).value =float(density)
    Sheet1.cell(row=j+2,column=10).value =viscos
    Sheet1.cell(row=j+2,column=11).value =SHC
    Sheet1.cell(row=j+2,column=12).value =TC_K
    Sheet1.cell(row=j+2,column=13).value =diffus #thermal diffusivity
    Sheet1.cell(row=j+2,column=14).value =delT
    Sheet1.cell(row=j+2,column=15).value =R
    Sheet1.cell(row=j+2,column=16).value =exp_fac
    Sheet1.cell(row=j+2,column=17).value =Gr
    Sheet1.cell(row=j+2,column=18).value =Pr
    Sheet1.cell(row=j+2,column=19).value =x    
    Sheet1.cell(row=j+2,column=20).value =y
    Sheet1.cell(row=j+2,column=21).value =x*y
    Sheet1.cell(row=j+2,column=22).value =x**2
    sx.append(x)
    sy.append(y)
    sxy.append(x*y)
    sx2.append(x**2)
    
    
    wb2.save ("cn.xlsx")
    #print(j+3)
    #time.sleep(10) # Sleep for 3 seconds
    
import math
sx_v=_sum(sx)
sy_v=_sum(sy)
sxy_v=_sum(sxy)
sx2_v=_sum(sx2)
N=7; #it debends upon good yield readings count
b=((N*sxy_v)-(sx_v*sy_v))/((N*sx2_v)-(sx_v**2))
#print (b)
a=(sy_v/N)-(b*(sx_v/N))
#print (a)
C=math.exp(a)
n=b
print("C and n values are :",C,n)
Sheet1.cell(row=9,column=19).value =sx_v
Sheet1.cell(row=9,column=20).value =sy_v
Sheet1.cell(row=9,column=21).value =sxy_v
Sheet1.cell(row=9,column=22).value =sx2_v  
Sheet1.cell(row=10,column=19).value ="C Value"
Sheet1.cell(row=10,column=20).value =C
Sheet1.cell(row=10,column=21).value ="n Value"
Sheet1.cell(row=10,column=22).value =n
wb2.save ("cn.xlsx")

