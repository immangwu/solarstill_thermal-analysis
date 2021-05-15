####----------------------------------------------------------------------------------------------------------------------------------------######
                                                        ##Radiation Related Values##
#Conventions taken from G N Tiwari boobk
#vARIATION OF ATTENUATION FACTOR WITH WATER DEPTH PG NO 281 IN TIWARI BOOK
def att_factor(depthw):
    switcher = {
        0.20: 0.51,
        0.10: 0.5492,
        0.08: 0.5648,
        0.06: 0.5858,
        0.04: 0.6185,
        0.02: 0.6756,
        0.01: 0.7344,
        0.008:0.7565,
        0.004:0.831,
        0.0  :1.0,
    }

    return switcher.get(depthw, 0)
#alpha'b alpha'w alpha'g
def alpha (alpha_w,alpha_g,alpha_b,R_g,R_w):
    att_fval = att_factor(depthw);
    alpha_gd = (1-R_g)*alpha_g;
    alpha_wd = (1-R_g)*(1-alpha_g)*(1-R_w)*(1-att_fval);
    alpha_bd = alpha_b*(1-R_g)*(1-alpha_g)*(1-R_w)*att_fval;
    return alpha_gd,alpha_wd,alpha_bd;

def latentheatofvapor(temp):
    L=(2.4935*10**6)*(1-((9.4779*10**-4)*temp)+((1.3132*10**-7)*temp**2)-((4.7974*10**-9)*temp**3));
    return L
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                    ##External Heat Transfer##
#Top loss Coefficient
def hrg(epsg,Tg,Ta):
    Tsky=Ta-6;
    hrg=(epsg*5.67*10**-8*abs((Tg+273)**4-(Tsky+273)**4))/abs(Tg-Ta);
    return hrg
def hcg (V):
    hcg = 2.8+3.0*V;
    return hcg

def h1g(epsg,Tg,Ta,V):
    hrg_v=hrg(epsg,Tg,Ta);
    hcg_v=hcg (V);
    h1g=hrg_v+hcg_v;
    return h1g

#Bottom and Side Loss Coefficient (bslc) K_i conductivity insulation L_i thickness of insulation
#bottom loss coefficient Ub V wind velocity m/sec
def Ub(V_basin,h_wbslc,K_i,L_i):
    h_common=5.7+(3.8*V)
    Ub = ((1/h_wbslc)+(L_i/K_i)+(1/h_common))**-1;
    return Ub
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                    ##Internal Heat Transfer##
#Conventions taken from G N Tiwari boobk
#the values of Pw and Pg
import math
def PT(Temp):
    PT = math.exp(25.317-(5144/(Temp+273)));
    return PT

#hrw is the radiative heat transfer coefficient from the water surface to the glass cover and is given by 
def hrw(eps_g,eps_w,T_w,T_g):
    eps_eff = ((1/eps_g)+(1/eps_w)-1)**-1
    hrw= eps_eff*5.67*10**-8*(((T_w+273)**2)+((T_g+273)**2))*(T_w+T_g+546);
    return hrw
#convective loss coefficient hcw
#dunkle model
def hcw_dunkle(T_w,T_g):
    P_w = PT(T_w)
    P_g = PT(T_g)
    hcw_dunkle = 0.884* ((abs(T_w-T_g)+((abs(P_w-P_g)*(T_w+273))/((268.9*10**3)-P_w)))**(1/3));
    return hcw_dunkle
#Evaporative loss coefficient hew

def hew(T_w,T_g):
##    if abs(T_w-T_g)==0:
##        hew=0
##    else:
    #print(T_w,T_g)
    P_w = PT(T_w)
    P_g = PT(T_g)
    hcw=hcw_dunkle(T_w,T_g)
    #print(hcw)
    hew=(16.273*10**-3)*hcw*(abs(P_w-P_g)/abs(T_w-T_g));
    #print(hew)
    return hew
def h1w(eps_g,eps_w,T_w,T_g):
    v_1=hrw(eps_g,eps_w,T_w,T_g)
    v_2=hcw_dunkle(T_w,T_g)
    v_3=hew(T_w,T_g)
    h1w=(v_1+v_2+v_3)
    return h1w
    
    ##h1w is total internal heat transfer coefficient
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                            ##Overall Heat transfer##

#Top Loss Coefficient

def Ut(h1g,h1w):
    Ut=((1/h1g)+(1/h1w))**-1;
    return Ut

def Ul(Ut_v,Ub_v):
    Ul = Ut_v+Ub_v;
    return Ul
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                            ###Effectiveness of the individual mode of heat transfer####
def fract(eps_g,eps_w,T_w,T_g):
    hcw_v=hcw_dunkle(T_w,T_g)
    hew_v=hew(T_w,T_g)
    hrw_v=hrw(eps_g,eps_w,T_w,T_g)
    h1w_v= h1w(eps_g,eps_w,T_w,T_g);
    qcw = hcw_v*abs(T_w-T_g);
    qew = hew_v*abs(T_w-T_g);
    qrw = hrw_v*abs(T_w-T_g);
    qwg = h1w_v*abs(T_w-T_g);
    F_ew = qew/qwg;
    F_cw = qcw/qwg;
    F_rw = qrw/qwg;
    return qcw,qew,qrw,qwg,F_ew,F_cw,F_rw



####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                ## yield and eff and exergy ##
def mew(T_w,T_g):
    hew_v=hew(T_w,T_g);
    L_v = latentheatofvapor(T_w)
    mew = ((hew_v*abs(T_w-T_g))/L_v)*3600
    return mew

def eff_ins(T_w,T_g,A_s,I_t):
    mew_v=mew(T_w,T_g);
    #print("mew is ",mew)
    L_v = (latentheatofvapor(T_w))
    #print("lv is",L_v)
    #print("area is" ,A_s)
    #print("solar Insolation is",I_t)
    eff_int = ((mew_v*L_v)/((I_t*3600)*A_s))*100;#1watt = 3600 j/hr
    return eff_int
def eff_jesus(eps_g,eps_w,T_w,T_g,Ta,V):
    h1w_v=h1w(eps_g,eps_w,T_w,T_g);
    h1g_v=h1g(eps_g,T_g,Ta,V);
    eff_int=((h1w_v*h1g_v)/(h1w_v+h1g_v))*abs(T_w-T_g)
    return eff_int

def exergy(T_w,T_g,A_s,Ta,I_t):
    T_s=5505;
    hew_v=hew(T_w,T_g);
    Exergy_evap=A_s*hew_v*abs(1-((Ta+273)/(T_w+273)));
    Exergy_in = A_s*I_t*abs(1-((4/3)*((Ta+273)/(T_s+273)))+((1/3)*(((Ta+273)/(T_s+273))**4)))
    eff_exergy = (Exergy_evap/Exergy_in)*100
    return eff_exergy


####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                            ###General variables###
#depthw = 0.10; #depth of water in m 10 cm = 0.01m
#alpha_w = 0.2 ; #water absorbtivity 0.2 from tiwari boobk
#alpha_g = 0.0 ; # absorbtivity of glass cover
#alpha_b = 0.075; # absorbtivity of basin cover
#R_g = 0.05; # reflectivity of glass cover
#R_w =0.05;  # reflectivity of water
eps_g = 0.94; #emisivity of glass cover
eps_w = 0.95 #emissivity of water
##T_w = 20;#water temperature
##T_g=22; #glass gover temperature
##Ta=8.5; #ambient temperature
#Tsky=Ta-6;
#V=3;
#V_basin =0; #basin wind velocity m/s zero always
h_wbslc=100; #W/m^2deg celcius taken from book need to find it
#K_i = 0.04; #Insulation thermal conductivity [W/m K]
#L_i = 0.5*10**-2; #Insulation thickness [m]
A_s=1;#surface area of the basin
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
##import array as arr                                                            ###OUTPUT###
####w_T = arr.array('f',[20,46.6,53.9,59.1,59.7,58.6,56.5,53.1]);
####g_T = arr.array('f',[12,37.5,42.3,46.2,46.3,44.4,43.5,38.9]);
####a_T = arr.array('f',[8.5,43,44,45,46,46,45,44]);
##w_T = arr.array('f',[14,28,44,55,63,62,61,60,55,41,33,30]);
##g_T = arr.array('f',[20,30,43,50,55,54,53,48,43,32,25,23,22]);
##I_t = arr.array('f',[650,900,1000,1010,980,800,600,400,90,10,9,8]);
##Ta  = arr.array('f',[24,28,32,34,33,27,24,22,20,18,17,16,15]);


##for i in range(12):
##    
##    print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
##    #print("Convective heat transfer coefficient is :," ,hcw_dunkle(w_T[i],g_T[i])  ,"[W/m2-K] ");
####
##for i in range(12):
##    
##    #print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
##    print("Convective heat transfer coefficient is :," ,hcw_dunkle(w_T[i],g_T[i])  ,"[W/m2-K] ");
##for i in range(12):
##    
##    #print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
##    print("Radiative heat transfer coefficient is :," ,hrw(eps_g,eps_w,w_T[i],g_T[i])  ,"[W/m2-K] ");
##
##for i in range(12):
##    
##    #print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
##    qcw,qew,qrw,qwg,F_ew,F_cw,F_rw=fract(eps_g,eps_w,w_T[i],g_T[i])
##    print("F_ew is",F_ew)
##    print("F_cw is",F_cw)
##    print("F_rw is",F_rw)
##    #print("qcw,qew,qrw,qwg,F_ew,F_cw,F_rw is :," ,fract(eps_g,eps_w,w_T[i],g_T[i])  ,"[W/m2-K] ");
##
##
##for i in range(12):
##    
##    #print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
##    print("Hourly productivity [kg/m2-h] is :," ,mew(w_T[i],g_T[i])  ,"[kg/m2-h] ");
    
##for i in range(12):
##    
##    #print("Evaporative heat transfer coefficient is :," ,hew(w_T[i],g_T[i])  ,"[W/m2-K] ");
####    print("Hourly productivity [kg/m2-h] is :," ,mew(w_T[i],g_T[i])  ,"[kg/m2-h] ");
####    print("solar Insolation is:",I_t[i]);
####    print("Latent heat is:",latentheatofvapor(w_T[i]));
##    print("Instantaneous Internal Efficiencies is :," ,eff_ins(w_T[i],g_T[i],A_s,I_t[i])  ,"% ");
##    #print("Instantaneous Internal Efficiencies theoretical is :," ,eff_jesus(eps_g,eps_w,w_T[i],g_T[i],Ta[i],V)  ,"% ");
##    print ("exergy is:",exergy(w_T[i],g_T[i],A_s,Ta[i],I_t[i]))
####----------------------------------------------------------------------------------------------------------------------------------------######
####----------------------------------------------------------------------------------------------------------------------------------------######
                                                        ##Excel##
import openpyxl
import xlrd
from pathlib import Path
import xlwt
from openpyxl import load_workbook


loc = ("1.xlsx");
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
wb2 = load_workbook (loc)
sheets=wb2.sheetnames
Sheet1=wb2 [sheets[0]]


for j in range(sheet.nrows-1):
#for i in range(sheet.ncols-3):
##    print(sheet.nrows)
##    print(sheet.ncols)
    time=(sheet.cell_value(j+1, 0))
    I_t= (sheet.cell_value(j+1, 1))
    Ta=(sheet.cell_value(j+1, 2))
    w_T=(sheet.cell_value(j+1, 3))
    g_T=(sheet.cell_value(j+1, 4))
    P_w_val=PT(w_T);
    P_g_val=PT(g_T);
    hew_val=hew(w_T,g_T);
    hcw_val=hcw_dunkle(w_T,g_T);
    hrw_val=hrw(eps_g,eps_w,w_T,g_T);
    h1w_val=h1w(eps_g,eps_w,w_T,g_T);
    qcw,qew,qrw,Qtotal_val,Few_val,Fcw_val,Frw_val=fract(eps_g,eps_w,w_T,g_T);
    yield_val=mew(w_T,g_T);
    Eff_val=eff_ins(w_T,g_T,A_s,I_t);
    Exergy_val= exergy(w_T,g_T,A_s,Ta,I_t)
    #print(time,solar_ins,Tempa,TempW,Tempg)
    Sheet1.cell(row=j+2,column=6).value =P_w_val
    Sheet1.cell(row=j+2,column=7).value =P_g_val
    Sheet1.cell(row=j+2,column=8).value =hew_val
    Sheet1.cell(row=j+2,column=9).value =hcw_val
    Sheet1.cell(row=j+2,column=10).value =hrw_val
    Sheet1.cell(row=j+2,column=11).value =h1w_val
    Sheet1.cell(row=j+2,column=12).value =Few_val
    Sheet1.cell(row=j+2,column=13).value =Fcw_val
    Sheet1.cell(row=j+2,column=14).value =Frw_val
    Sheet1.cell(row=j+2,column=15).value =Qtotal_val
    Sheet1.cell(row=j+2,column=16).value =yield_val
    Sheet1.cell(row=j+2,column=17).value =Eff_val
    Sheet1.cell(row=j+2,column=18).value =Exergy_val
    
wb2.save ("1.xlsx")

